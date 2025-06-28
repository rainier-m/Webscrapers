#!python3
# -*- coding: utf-8 -*-
'''
Created on May 23, 2023
Modified on May 30, 2023
Version 0.01.b
@author: rainier.madruga@gmail.com
A simple Python Program to push an Excel file to a database for monitoring Comics subscriptions
### =========================================================================================== ###
### Change & Revision Log                                                                       ###
### Date             Dev        Change Description                                              ###
### =========================================================================================== ###
    2023-05May-23    RWN        Initial Creation of this Python script
    2023-05May-30    RWM        Update to parse out Product Details and Title Info
'''

scriptVer = '0.01.c'

# Import necessary Libraries
import os
import datetime
import sys
import codecs
import pandas as pd
import mysql.connector
import openpyxl
import re

# Provide Google Drive Physical Location
drive_loc = 'D:\\OneDrive - Mdga, Inc\\Personal\\Comics\\'
file_name = 'MasterList_Comics_20240408_0904_04102024.xlsx'

wb = openpyxl.load_workbook(filename = drive_loc + file_name)
sheet = wb['Base']
wb_max_column = sheet.max_column
wb_max_row = sheet.max_row

# Establish the process Date & Time Stamp
ts = datetime.datetime.now().strftime("%H:%M:%S")
ds = datetime.datetime.now().strftime("%Y-%m-%d")
date = datetime.datetime.now().strftime("%Y%m%d")

# Updates the Time Stamp
def updateTS():
    update = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return update

# Set Character Output
print('System Encoding:', sys.stdout.encoding)
sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())

# Establish MySQL Connection
cnx = mysql.connector.connect(user='root',
                              password='password',
                              host='dewalt',
                              database='comics',
                              port='3306',
                              auth_plugin='mysql_native_password')

if cnx.is_connected():
    db_Info = cnx.get_server_info()
    print("Successfully connected to DB", db_Info)

cursor = cnx.cursor()

# Create Pandas dataframe for Google Drive inventory
pd.set_option("display.max_rows", None, "display.max_columns", None)

hr = " >>> *** ======================================================================= *** <<<"
shr = " >>> *** ==================== *** <<<"

# print (wb_max_row, wb_max_column)
print (hr)

def valid_date(value):
    date_input = value
    if (isinstance(date_input, datetime.date)) == True:
        date_output = date_input.strftime("%Y-%m-%d")
    else:
        date_output = '1900-01-01'
    return date_output

def pubStatus(value):
    data_input = value
    if (isinstance(data_input, datetime.date)) == True:
        status = 'Active'
    else:
        status =  data_input
    return status

def return_pub(value):
    match value:
        case 'Marvel':
            pub_out = 1
        case 'DC':
            pub_out = 2
        case 'Dark Horse':
            pub_out = 3
        case 'Dynamite':
            pub_out = 4
        case 'IDW':
            pub_out = 5
        case 'Image':
            pub_out = 6
        case 'Oni Press':
            pub_out = 7
        case 'Titan':
            pub_out = 8
        case 'Valiant':
            pub_out = 9
        case 'Boom!':
            pub_out = 10
        case _:
            pub_out = 99

    return pub_out

def get_product(value):
    input = value
    product_value = ''
    # \ [HC]
    tradePaperback = re.search("\sTP|\sSC", input)
    graphicNovel = re.search("\sGN", input)
    tradeIssue = re.search("#\d{1,4}", input)
    hardCover = re.search("\sHC", input)

    if tradeIssue != None:
        product_value = 'Issue'
    elif graphicNovel != None:
        product_value = 'GN'
    elif tradePaperback != None:
        product_value = 'TP'
    elif hardCover != None:
        product_value = 'HC'
    else:
        product_value = 'Merch'

    return product_value

# Take Base Excel and Process Data
for row in range(2, sheet.max_row+1):
    # Map in values from the read Excel
    current_publish_date = sheet['A' + str(row)].value
    original_publish = sheet['B' + str(row)].value
    publisher = sheet['C' + str(row)].value
    title = sheet['D' + str(row)].value
    run_date = sheet['E' + str(row)].value

    # Short Row for Title
    print(shr)

    # Get Date Values for String Dates
    current_date = valid_date(current_publish_date)
    original_publish_date = original_publish.strftime("%Y-%m-%d")

    if original_publish == None:
        original_publish = ''

    # Get Publisher values
    insert_pub = return_pub(publisher)
    # print (title)

    # Get Product Type and Status for Issues
    getProduct = get_product(title)
    status = pubStatus(current_publish_date)

    # Sets up the parsed out file
    titleName = ''
    titleDescr = ''
    issueDescr = ''

    # print(title)
    # For Comic Book Issues
    if getProduct == "Issue":
        findIssue = re.search("#\d{1,4}", title)
        # print (title)

        # Provides the Comic Issue Title
        titleName = title[0:findIssue.start()-1]
        # print(titleName)

        # Provides the Issue Number
        issueDescr = title[findIssue.start()+1:findIssue.end()]
        # print (title[findIssue.start():findIssue.end()+1])
        # print (findIssue.end(), len(title))

        # Provides the Extra Issue Information
        if len(title) > findIssue.end()+1:
            titleDescr = title[findIssue.end()+1:]
            # print (title[findIssue.end():])
        # print (shr)

    # For Comic Book Trade Paperbacks
    # print (getProduct)
    if getProduct == 'TP' or getProduct == 'HC' or getProduct == 'GN':
        # print (title)
        findTitle = re.search("\sGN|\sTP|\sHC|\sSC", title)
        titleName = title[0:findTitle.start()]

        # Provides the Extra Description Information, if any
        if len(title) > findTitle.end() + 1:
            titleDescr = title[findTitle.end()+1:]
    # For Merchandise
    if getProduct == 'Merch':
        titleName = title
        print (title)

    # print (status, insert_pub, getProduct, publisher, current_date, original_publish_date, title)
    # print (titleName, "|", getProduct, "|", titleDescr, "|", issueDescr, "|", publisher, "|", current_date, "|", status)

    print(title)
    print (current_date, "|", original_publish_date, "|", publisher, "|", insert_pub, "|", titleName, "|", status, "|", getProduct, "|", titleDescr, "|", issueDescr)
    # print (current_date, original_publish_date, publisher, title, run_date.strftime("%Y-%m-%d"))

    # Remove SQL Break Characters
    index = title.find('\'')
    if index != -1:
        title = title.replace('\'', ' ')
    # print (title)

    indexDescr = titleDescr.find('\'')
    if indexDescr != -1:
        titleDescr = titleDescr.replace('\'', ' ')
    # print (titleDescr)

    # Check to see if in Raw Table
    checkReleaseItem = "SELECT * FROM raw_comics_publish_file where current_publish_date = '%s' and original_publish_date = '%s' and publisher = '%s' and title = '%s' and file_rundate = '%s'" % (current_date, original_publish_date, publisher, title, run_date.strftime("%Y-%m-%d"))

    # print(checkReleaseItem)

    cursor.execute(checkReleaseItem)
    releaseCheck = cursor.fetchone()

    if releaseCheck == None:
        insertReleaseItem = "INSERT INTO raw_comics_publish_file (current_publish_date, original_publish_date, publisher, title, file_rundate, status, productType, titleDesc, issueDescr) VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % (current_date, original_publish_date, publisher, title, run_date.strftime("%Y-%m-%d"), status, getProduct, titleDescr, issueDescr)
        cursor.execute(insertReleaseItem)
        cnx.commit()
    else:
        print ("Already published", title, run_date)


